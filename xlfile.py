import openpyxl
import pandas as pd

class ExcelFile:

    def __init__(self, file_name):
        self.filename = file_name
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook.active

    def close_file(self):
        self.workbook.close()

    def countRows(self):
        # Count rows
        return self.worksheet.max_row
    
    def countColumns(self):
        # Count columns
        return self.worksheet.max_column

    def writeOnXL(self, data):

        self.update_excel(self.countRows() + 1, data)

    # ---> FOR SO_FORM <--- #
    def read_into_dataframe_SO(self):
        # Read the Excel file into a pandas DataFrame
        return pd.read_excel(self.filename,  dtype={'PHONE': str, 'ZIPCODE': str})

    # ---> FOR AI_FORM <--- #
    def read_into_dataframe(self):
        # Read the Excel file into a pandas DataFrame
        return pd.read_excel(self.filename, 'Current Master')
    
    def update_excel(self, row, new_values):
        # Update the Excel file with the new values
        for i, value in enumerate(new_values, start=1):
            self.worksheet.append(row=row, column=i, value=value)
        self.workbook.save(self.filename)
        self.close_file()

    def empty_row(self):
        # Empty a row at the end of the excel
        row = self.countRows() + 1
        for i in range(1, self.countColumns() + 1):
            self.worksheet.cell(row=row, column=i).value = " "
        self.workbook.save(self.filename)
        self.close_file()

    def add_row(self, dict, row_index=0):
        row_index=self.countRows() + 1
        df = pd.DataFrame([dict])
        existing_df = self.read_into_dataframe_SO()
        updated_df = pd.concat([existing_df.iloc[:row_index], df, existing_df.iloc[row_index:]], ignore_index=True)
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            updated_df.to_excel(writer, sheet_name=self.worksheet.title, index=False)
        self.close_file()



def open_excel_file(file_name): 
    return ExcelFile(file_name)
